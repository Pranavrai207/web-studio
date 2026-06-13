"""
================================================================================
 scraper.py - Delhi Local Business Lead Generation Scraper
================================================================================
 Tool Stack : Playwright (Chromium) + pandas + openpyxl
 Author     : Antigravity
 Purpose    : Scrape Google Maps for local business leads in Delhi, save to Excel
 Usage      : python scraper.py
              or: python scraper.py --query "restaurants in Delhi" --max 40
================================================================================
"""

import argparse
import logging
import random
import sys
import time
from typing import Optional

import pandas as pd
from playwright.sync_api import Page, Playwright, sync_playwright, TimeoutError as PlaywrightTimeout
from playwright_stealth import Stealth

import config
import db_manager

# ─────────────────────────────────────────────────────────────────────────────
#  LOGGING SETUP
# ─────────────────────────────────────────────────────────────────────────────

def setup_logging() -> logging.Logger:
    """Configure root logger with console + optional file handler."""
    log_level = getattr(logging, config.LOG_LEVEL.upper(), logging.INFO)
    fmt = "%(asctime)s [%(levelname)s] %(message)s"
    handlers: list[logging.Handler] = [logging.StreamHandler(sys.stdout)]

    if config.LOG_FILE:
        handlers.append(logging.FileHandler(config.LOG_FILE, encoding="utf-8"))

    logging.basicConfig(level=log_level, format=fmt, handlers=handlers)
    return logging.getLogger(__name__)


logger = setup_logging()


# ─────────────────────────────────────────────────────────────────────────────
#  UTILITY: Human-like delays
# ─────────────────────────────────────────────────────────────────────────────

def human_delay(min_s: float = None, max_s: float = None) -> None:
    """Sleep for a random duration to mimic human browsing behaviour."""
    lo = min_s if min_s is not None else config.MIN_DELAY
    hi = max_s if max_s is not None else config.MAX_DELAY
    duration = random.uniform(lo, hi)
    logger.debug("Sleeping %.2fs ...", duration)
    time.sleep(duration)


# ─────────────────────────────────────────────────────────────────────────────
#  1. initialize_browser()
# ─────────────────────────────────────────────────────────────────────────────

def initialize_browser(playwright: Playwright):
    """
    Launch a Chromium browser with realistic viewport and user-agent settings
    along with stealth evasions to bypass bot detection.

    Returns:
        tuple[Browser, BrowserContext, Page]
    """
    logger.info("Launching Chromium browser ...")

    # Select random user-agent
    ua_list = getattr(config, 'USER_AGENTS', [])
    user_agent = random.choice(ua_list) if ua_list else (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    )
    logger.info("Using randomized User-Agent: %s", user_agent)

    # Randomized viewport size
    viewport_width = random.randint(1366, 1600)
    viewport_height = random.randint(768, 900)
    logger.info("Using randomized Viewport: %dx%d", viewport_width, viewport_height)

    browser = playwright.chromium.launch(
        headless=config.HEADLESS,
        slow_mo=config.BROWSER_SLOW_MO,
        args=[
            "--disable-blink-features=AutomationControlled",  # hide WebDriver flag
            "--no-sandbox",
            "--disable-dev-shm-usage",
            "--disable-gpu",
        ],
    )

    # Realistic desktop context
    context = browser.new_context(
        viewport={"width": viewport_width, "height": viewport_height},
        user_agent=user_agent,
        locale="en-IN",
        timezone_id="Asia/Kolkata",
        geolocation={"latitude": 28.6139, "longitude": 77.2090},  # New Delhi
        permissions=["geolocation"],
        device_scale_factor=random.choice([1.0, 1.25, 1.5]),
        has_touch=random.choice([True, False]),
    )

    # Mask navigator.webdriver property
    context.add_init_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )

    page = context.new_page()

    # Apply playwright-stealth
    if getattr(config, 'STEALTH_ENABLED', True):
        try:
            Stealth().apply_stealth_sync(page)
            logger.info("Playwright stealth applied successfully.")
        except Exception as e:
            logger.warning("Failed to apply playwright-stealth: %s", e)

    logger.info("Browser ready.")
    return browser, context, page


# ─────────────────────────────────────────────────────────────────────────────
#  2. perform_search()
# ─────────────────────────────────────────────────────────────────────────────

def perform_search(page: Page, query: str) -> bool:
    """
    Navigate to Google Maps and enter the search query.
    Waits for the results panel to appear.

    Key fix: use 'domcontentloaded' NOT 'networkidle' - Google Maps streams
    map tiles indefinitely so networkidle never fires, causing a 30-second hang.

    Args:
        page  : Active Playwright page.
        query : Search string, e.g. "cafes in Delhi".

    Returns:
        True if results loaded, False on failure.
    """
    from urllib.parse import quote_plus
    search_url = f"https://www.google.com/maps/search/{quote_plus(query)}"
    logger.info("Navigating to: %s", search_url)

    try:
        # Use 'domcontentloaded' - Maps never reaches 'networkidle'
        page.goto(search_url, wait_until="domcontentloaded", timeout=45_000)
        human_delay(3, 5)

        # Accept cookies / consent dialog if it appears
        for consent_sel in [
            'button:has-text("Accept all")',
            'button:has-text("I agree")',
            'button[aria-label="Accept all"]',
            'form:has(button) button',
        ]:
            try:
                btn = page.locator(consent_sel).first
                if btn.is_visible(timeout=3_000):
                    btn.click()
                    logger.info("Dismissed consent dialog.")
                    human_delay(2, 3)
                    break
            except Exception:
                pass

        # Wait for the results feed - try multiple known selectors
        FEED_SELECTORS = [
            'div[role="feed"]',
            'div[aria-label*="Results"]',
            '[jstcache] div[tabindex="0"] div[role="article"]',
        ]
        feed_found = False
        for sel in FEED_SELECTORS:
            try:
                page.wait_for_selector(sel, timeout=25_000)
                logger.info("Results panel loaded (selector: %s)", sel)
                feed_found = True
                break
            except PlaywrightTimeout:
                logger.debug("Selector not found: %s - trying next ...", sel)
                continue

        if not feed_found:
            logger.error("Could not locate results panel with any known selector.")
            return False

        return True

    except PlaywrightTimeout:
        logger.error("Page navigation timed out (45s).")
        return False
    except Exception as exc:
        logger.error("Error during search: %s", exc)
        return False


# ─────────────────────────────────────────────────────────────────────────────
#  3. scroll_results()
# ─────────────────────────────────────────────────────────────────────────────

def scroll_results(page: Page) -> None:
    """
    Gradually scroll the results panel to load more listings.
    Uses incremental, randomised scroll amounts to mimic human behaviour.
    """
    logger.info("Scrolling results panel ...")

    try:
        # Locate the scrollable feed panel
        panel = page.locator('div[role="feed"]').first
        panel.hover()

        for step in range(config.SCROLL_STEPS):
            # Vary scroll amount slightly each step
            amount = config.SCROLL_AMOUNT + random.randint(-80, 80)
            page.mouse.wheel(0, amount)
            logger.debug("Scroll step %d/%d (+%dpx)", step + 1, config.SCROLL_STEPS, amount)
            human_delay(config.SCROLL_DELAY_MIN, config.SCROLL_DELAY_MAX)

    except Exception as exc:
        logger.warning("Scroll interrupted: %s", exc)


# ─────────────────────────────────────────────────────────────────────────────
#  4. extract_business_data()
# ─────────────────────────────────────────────────────────────────────────────

def _is_platform_website(url: str) -> bool:
    """Check if the URL belongs to a known platform (Zomato, Swiggy, etc.)."""
    if not url or url == "N/A":
        return False
    url_lower = url.lower()
    return any(domain in url_lower for domain in getattr(config, 'PLATFORM_DOMAINS', []))


def _score_lead(website: str, rating: Optional[float]) -> str:
    """
    Assign a lead priority score based on strict rules:
    - No website link at all ("N/A") -> "[HIGH] High"
    - Platform link + High rating   -> "[MEDIUM] Medium"
    - Platform link + Low rating    -> "[LOW] Low"
    """
    if not website or website == "N/A":
        return "[HIGH] High"
    
    if _is_platform_website(website):
        if rating and rating >= config.HIGH_RATING_THRESHOLD:
            return "[MEDIUM] Medium"
        return "[LOW] Low"
    
    # This point should not be reached due to filtering, but return Low as fallback
    return "[LOW] Low"


def _is_local_business(name: str, reviews: Optional[int]) -> bool:
    """
    Return True if the listing looks like a small/local business.
    Return False (skip it) if it matches a known big brand or mall.

    Rules:
      1. Name must NOT contain any word from BRAND_BLACKLIST.
      2. Review count must be below MAX_REVIEWS_FOR_LOCAL (big chains
         almost always have thousands of reviews).
    """
    name_lower = name.lower()

    # Check against brand blacklist
    blacklist = getattr(config, 'BRAND_BLACKLIST', set())
    for brand in blacklist:
        if brand.lower() in name_lower:
            logger.debug("Brand filtered out: %s (matched: '%s')", name, brand)
            return False

    # Check review count cap
    max_reviews = getattr(config, 'MAX_REVIEWS_FOR_LOCAL', 2000)
    if reviews and reviews > max_reviews:
        logger.debug("High-review filtered: %s (%d reviews)", name, reviews)
        return False

    return True


def extract_business_data(page: Page, max_results: int) -> list[dict]:
    """
    Iterate over visible listing cards and extract structured data from each.

    Strategy:
      1. Collect all listing card locators.
      2. Click each card to open its detail panel.
      3. Parse detail panel for all required fields.
      4. De-duplicate by (name, address) composite key.

    Args:
        page        : Active Playwright page.
        max_results : Stop collecting after this many unique records.

    Returns:
        List of business dicts.
    """
    results: list[dict] = []
    seen_keys: set[str] = set()  # deduplication set

    logger.info("Extracting business data (target: %d) ...", max_results)

    # ── Scroll loop: keep scrolling + extracting until we hit max_results ──
    scroll_cycles = 0
    max_scroll_cycles = getattr(config, 'MAX_SCROLL_CYCLES', 25)  # from config

    while len(results) < max_results and scroll_cycles < max_scroll_cycles:
        scroll_cycles += 1

        # Get all currently loaded listing cards
        cards = page.locator('div[role="feed"] > div > div[jsaction]').all()
        logger.info("Cycle %d: found %d cards, collected %d so far",
                    scroll_cycles, len(cards), len(results))

        for card in cards:
            if len(results) >= max_results:
                break

            business = _extract_single_card(page, card)
            if business is None:
                continue

            # ── Local business filter (skip chains & malls) ──────────────
            if not _is_local_business(
                business['Name'],
                business.get('Total Reviews') if isinstance(business.get('Total Reviews'), int) else None
            ):
                continue

            # ── Personal Website Classification & Database Sync ──────────
            website = business.get('Website', 'N/A')
            is_personal_site = (website != "N/A" and not _is_platform_website(website))
            
            if is_personal_site:
                # Store business with website as competitor in DB
                business["status"] = "has_website"
                db_manager.insert_lead(business)
                logger.debug("Competitor website saved to SQLite: %s (%s)", business["Name"], website)
                continue
            
            # Lead is a prospect (no personal website)
            business["status"] = "scraped"
            
            # Find a local competitor with a website for matching
            comp_name, comp_web = db_manager.find_competitor_for_lead(business["Category"], business["Address"])
            business["Competitor Name"] = comp_name
            business["Competitor Website"] = comp_web
            
            # Save prospect to SQLite
            db_manager.insert_lead(business)

            # ── Phone Number Requirement Filter ──────────────────────────
            # Skip if business has no phone number (if required in config)
            if getattr(config, 'REQUIRE_PHONE_NUMBER', False):
                phone = business.get('Phone', 'N/A')
                if not phone or phone == "N/A":
                    logger.debug("No phone number found - skipping: %s", business["Name"])
                    continue

            # Deduplication check
            dedup_key = f"{business['Name'].lower()}|{business['Address'].lower()}"
            if dedup_key in seen_keys:
                logger.debug("Duplicate skipped: %s", business["Name"])
                continue

            seen_keys.add(dedup_key)
            results.append(business)
            logger.info("[%d/%d] Added prospect: %s (Rival: %s)", len(results), max_results, business["Name"], comp_name)

        if len(results) < max_results:
            scroll_results(page)  # load more listings
            human_delay(2, 4)

    logger.info("Extraction complete. Total unique records: %d", len(results))
    return results


def _has_chain_signal(page: Page) -> bool:
    """
    Detect if the currently-open Google Maps detail panel belongs to a
    chain or franchise business, using Google Maps' own UI signals.

    Signals checked (all visible in the detail side-panel):
      1. 'All locations' / 'See all locations' link  → multiple outlets = chain
      2. 'Part of' label (e.g. "Part of McDonald's")
      3. 'More from this business' / 'Locations' section header
      4. aria-labels or href containing "place_id" list (multi-location)
      5. 'Owned by' or 'Subsidiary of' text in about section
    """
    chain_text_signals = [
        "all locations",
        "see all locations",
        "see more locations",
        "other locations",
        "more locations",
        "part of ",
        "owned by",
        "subsidiary of",
        "franchise",
        "chain store",
    ]

    # Check 1: Presence of an 'all locations' type link/button
    for selector in [
        'a[href*="/maps/place"][aria-label*="location"]',
        'button[aria-label*="location"]',
        '[data-value*="locations"]',
        'a[jsaction*="alllocations"]',
    ]:
        try:
            if page.locator(selector).count() > 0:
                logger.debug("Chain signal: locations link found")
                return True
        except Exception:
            pass

    # Check 2: Scan visible text for chain keywords
    try:
        panel_text = page.locator(
            'div[role="main"], div[jsaction*="pane"]'
        ).first.inner_text(timeout=2_000).lower()

        for signal in chain_text_signals:
            if signal in panel_text:
                logger.debug("Chain signal: text '%s' found", signal)
                return True
    except Exception:
        pass

    return False


def _extract_single_card(page: Page, card) -> Optional[dict]:
    """
    Click a listing card and scrape all fields from its detail side-panel.

    Returns a dict on success, or None if the card should be skipped.
    """
    try:
        # Scroll card into view and click it
        card.scroll_into_view_if_needed(timeout=5_000)
        card.click(timeout=5_000)
        human_delay(2, 4)

        # Wait for the detail pane to settle
        page.wait_for_selector('h1.DUwDvf, h1[class*="fontHeadlineLarge"]',
                               timeout=10_000)

        # ── Name ──────────────────────────────────────────────────────────
        name = _safe_text(page,
            'h1.DUwDvf, h1[class*="fontHeadlineLarge"]')
        if not name:
            return None  # not a valid business card

        # ── Chain / Franchise Signal Check ───────────────────────────────
        # Ask Google Maps itself if this is a chain - before scraping anything
        if _has_chain_signal(page):
            logger.debug("Chain detected via Maps signal: %s - skipping", name)
            return None

        # ── Rating ────────────────────────────────────────────────────────
        # Try multiple selector patterns - Google Maps changes class names often
        rating_str = ""
        for r_sel in [
            'span.ceNzKf',
            'span[aria-hidden="true"].MW4etd',
            'div.F7nice span[aria-hidden="true"]',
            'span[role="img"][aria-label*="star"]',
            'div[jsaction*="rating"] span[aria-hidden]',
        ]:
            rating_str = _safe_text(page, r_sel)
            if rating_str and rating_str.replace(".", "").replace(",", "").strip().isdigit():
                break
            # Also try aria-label which often has "Rated X.X stars"
            val = _safe_attr(page, 'span[aria-label*="star"]', 'aria-label')
            if val:
                import re as _re
                m = _re.search(r'[\d]+[.,][\d]', val)
                if m:
                    rating_str = m.group()
                    break

        rating: Optional[float] = None
        if rating_str:
            try:
                rating = float(rating_str.replace(",", ".").strip())
                if not (0.0 <= rating <= 5.0):
                    rating = None  # sanity check
            except ValueError:
                pass

        # ── Total Reviews ─────────────────────────────────────────────────
        reviews_str = _safe_text(page,
            'div[jsaction*="pane.rating"] span[aria-label*="review"], '
            'button[jsaction*="pane.review"] span, '
            'span[aria-label*="reviews"]')
        reviews = _parse_reviews(reviews_str)
        # ── Ultimate Gatekeeper: Review Count ─────────────────────────────
        if reviews and reviews > config.MAX_REVIEWS_FOR_LOCAL:
            logger.debug("High reviews (%d) detected (Gatekeeper) - skipping %s", reviews, name)
            return None

        # ── Address ───────────────────────────────────────────────────────
        address = _safe_text(page,
            'button[data-item-id="address"] div.fontBodyMedium, '
            'div[data-item-id="address"] .rogA2c .Io6YTe, '
            'button[data-tooltip="Copy address"] .fontBodyMedium')

        # ── Phone Number ──────────────────────────────────────────────────
        phone = _safe_text(page,
            'button[data-item-id*="phone"] div.fontBodyMedium, '
            'a[href^="tel:"] div.fontBodyMedium')
        if not phone:
            # Fallback: look for tel: href directly
            tel_link = page.locator('a[href^="tel:"]').first
            if tel_link.count() > 0:
                href = tel_link.get_attribute("href") or ""
                phone = href.replace("tel:", "").strip()

        # ── Website URL ───────────────────────────────────────────────────
        # Must start with http/https - reject Maps UI text like "Add website"
        website = ""
        for w_sel in [
            'a[data-item-id="authority"]',
            'a[aria-label*="website"]',
            'a[href^="http"]:not([href*="google"])',
        ]:
            raw = _safe_attr(page, w_sel, "href")
            if raw and raw.startswith("http"):
                website = raw
                break

        # ── Check B: Website franchise signal ─────────────────────────────
        if _is_franchise_website(website):
            logger.debug("Franchise website detected: %s - skipping %s", website, name)
            return None

        # ── Check C: Address franchise signal ─────────────────────────────
        # Logic-3 (Refined): 
        # - High confidence (Malls/Airports) -> Skip immediately
        # - Soft signals (Complex/Society) -> Skip ONLY if reviews > 1000
        is_high_conf, is_soft = _check_address_signals(address)
        if is_high_conf:
            logger.debug("High-confidence chain address detected: %s - skipping %s", address, name)
            return None
        # (Soft signal review check is now handled by the global Gatekeeper above)

        # ── Category ─────────────────────────────────────────────────────
        category = _safe_text(page,
            'button[jsaction*="category"], '
            'span.DkEaL, '
            '.LBgpqf .fontBodyMedium button')

        # ── Google Maps URL ───────────────────────────────────────────────
        maps_url = page.url

        # ── Lead Priority ─────────────────────────────────────────────────
        priority = _score_lead(website, rating)

        return {
            "Name"               : name or "N/A",
            "Rating"             : rating if rating else "N/A",
            "Total Reviews"      : reviews or "N/A",
            "Address"            : address or "N/A",
            "Phone"              : phone or "N/A",
            "Website"            : website or "N/A",
            "Category"           : category or "N/A",
            "Maps URL"           : maps_url,
            "Priority"           : priority,
            "Competitor Name"    : "N/A",
            "Competitor Website" : "N/A",
        }

    except PlaywrightTimeout:
        logger.warning("Timeout opening listing card - skipping.")
        return None
    except Exception as exc:
        logger.warning("Failed to extract card data: %s", exc)
        return None


# ─── Franchise URL & Address Heuristics ─────────────────────────────────────

def _is_franchise_website(url: str) -> bool:
    """
    Check B: Franchise websites almost always reveal themselves via URL
    or page title keywords - no need to visit the page.

    Franchise signals in URL:
      - /franchise  /own-a-franchise  /become-a-partner
      - /locations  /find-a-store  /store-locator
      - /outlets    /branches
    """
    if not url or not url.startswith("http"):
        return False

    url_lower = url.lower()
    franchise_url_signals = [
        "franchise", "own-a", "become-a-partner", "become-a-franchis",
        "store-locator", "find-a-store", "find-a-branch",
        "/locations", "/outlets", "/branches", "/stores",
        "franchisee", "partner-with-us",
    ]
    return any(sig in url_lower for sig in franchise_url_signals)


def _check_address_signals(address: str) -> tuple[bool, bool]:
    """
    Check C: Franchise/chain outlets cluster in malls and large commercial
    complexes. Real local businesses typically have simple street addresses.

    Returns (is_high_confidence, is_soft_signal)
    """
    if not address:
        return False, False

    addr_lower = address.lower()

    # High-confidence chain address keywords (Malls, Terminals, Airports)
    high_conf_signals = [
        " mall", "select citywalk", "ambience mall", "dlf ",
        "vegas mall", "pacific mall", "unity one",
        "terminal ", "airport", "t1 ", "t2 ", "t3 ",
        "metro station concourse", "metro mall",
    ]
    
    # Soft signals: local shops can be here, but high-review ones are likely chains
    soft_signals = ["complex", "society", "shopping center", "mall road"]

    is_high = any(sig in addr_lower for sig in high_conf_signals)
    is_soft = any(sig in addr_lower for sig in soft_signals)

    return is_high, is_soft


# ─── Selector helpers ────────────────────────────────────────────────────────

def _safe_text(page: Page, selector: str) -> str:
    """Return inner text for the first matching element, or empty string."""
    try:
        el = page.locator(selector).first
        if el.count() > 0:
            return (el.inner_text(timeout=3_000) or "").strip()
    except Exception:
        pass
    return ""


def _safe_attr(page: Page, selector: str, attr: str) -> str:
    """Return an attribute value for the first matching element, or empty string."""
    try:
        el = page.locator(selector).first
        if el.count() > 0:
            return (el.get_attribute(attr, timeout=3_000) or "").strip()
    except Exception:
        pass
    return ""


def _parse_reviews(text: str) -> Optional[int]:
    """Extract a numeric review count from strings like '(1,234)' or '1234 reviews'."""
    import re
    if not text:
        return None
    text = text.replace(",", "").replace("(", "").replace(")", "")
    match = re.search(r"\d+", text)
    return int(match.group()) if match else None


# ─────────────────────────────────────────────────────────────────────────────
#  5. save_to_excel()
# ─────────────────────────────────────────────────────────────────────────────

def save_to_excel(data: list[dict], filepath: str) -> None:
    """
    Convert the collected records to a pandas DataFrame and write to .xlsx
    with auto-sized columns and a styled header row.

    Args:
        data     : List of business dicts.
        filepath : Destination .xlsx path.
    """
    if not data:
        logger.warning("No data to save.")
        return

    # Final pass to improve competitor matching using all newly scraped data
    logger.info("Improving competitor matches with newly scraped data...")
    for r in data:
        comp_name, comp_web = db_manager.find_competitor_for_lead(r.get("Category", "N/A"), r.get("Address", "N/A"))
        if comp_name != "N/A":
            r["Competitor Name"] = comp_name
            r["Competitor Website"] = comp_web
            db_manager.update_lead_competitor_info(r["Name"], r["Address"], comp_name, comp_web)

    df = pd.DataFrame(data, columns=[
        "Name", "Rating", "Total Reviews",
        "Address", "Phone", "Website",
        "Category", "Maps URL", "Priority",
        "Competitor Name", "Competitor Website"
    ])

    # ── Priority Sorting ─────────────────────────────────────────────
    # Sort: High (1) -> Medium (2) -> Low (3)
    priority_map = {"[HIGH] High": 1, "[MEDIUM] Medium": 2, "[LOW] Low": 3}

    # ── Load existing data if file exists ─────────────────────────────
    import os
    if os.path.exists(filepath):
        try:
            df_old = pd.read_excel(filepath)
            # Merge new with old, drop duplicates based on Name and Address
            df = pd.concat([df_old, df], ignore_index=True)
            df.drop_duplicates(subset=["Name", "Address"], keep="first", inplace=True)
            logger.info("Merged %d new leads with existing data.", len(data))
        except Exception as e:
            logger.warning("Could not merge with existing file: %s", e)

    df["_priority_sort"] = df["Priority"].map(priority_map).fillna(99)
    df = df.sort_values(by=["_priority_sort", "Name"]).drop(columns=["_priority_sort"])

    logger.info("Saving %d total unique records to %s ...", len(df), filepath)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Delhi Leads")

        ws = writer.sheets["Delhi Leads"]

        # ── Auto-size columns ──────────────────────────────────────────
        for col_cells in ws.columns:
            max_len = max(
                len(str(cell.value or "")) for cell in col_cells
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

        # ── Style header row ───────────────────────────────────────────
        from openpyxl.styles import Font, PatternFill, Alignment

        header_fill = PatternFill("solid", fgColor="1F4E79")   # dark navy
        header_font = Font(color="FFFFFF", bold=True, size=11)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Freeze top row
        ws.freeze_panes = "A2"

    logger.info("[SUCCESS] Excel saved: %s", filepath)


# ─────────────────────────────────────────────────────────────────────────────
#  6. main()
# ─────────────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    """Allow CLI overrides of the default config values."""
    parser = argparse.ArgumentParser(
        description="Delhi Local Business Lead Generator (Google Maps + Playwright)"
    )
    parser.add_argument(
        "--query", "-q",
        default=None,
        help="Search query, e.g. 'restaurants in Delhi'. "
             "Defaults to queries defined in config.py.",
    )
    parser.add_argument(
        "--max", "-m",
        type=int,
        default=config.MAX_RESULTS_PER_QUERY,
        help=f"Maximum results per query (default: {config.MAX_RESULTS_PER_QUERY})",
    )
    parser.add_argument(
        "--output", "-o",
        default=config.OUTPUT_FILE,
        help=f"Output Excel filename (default: {config.OUTPUT_FILE})",
    )
    return parser.parse_args()


def main() -> None:
    """
    Orchestrate the entire lead-generation pipeline:
      1. Parse CLI args
      2. Launch browser
      3. For each search query:
          a. perform_search()
          b. extract_business_data()
      4. Aggregate, deduplicate, save to Excel
      5. Clean shutdown
    """
    args = parse_args()

    # Determine which queries to run
    queries = [args.query] if args.query else config.SEARCH_QUERIES

    all_records: list[dict] = []
    seen_global: set[str] = set()

    # ── Initialize SQLite database ──
    db_manager.init_db()

    # ── Load existing leads to seen_global for persistent deduplication ──
    # 1. From SQLite
    try:
        # Load excluded leads (qualified, deal_cracked, disqualified)
        excluded_keys = db_manager.load_excluded_leads()
        seen_global.update(excluded_keys)
        logger.info("Loaded %d excluded leads from SQLite database (will be skipped).", len(excluded_keys))
        
        # Load all other leads from SQLite database to prevent duplicate scraping
        with db_manager.get_connection() as conn:
            cur = conn.execute("SELECT name, address FROM leads")
            for row in cur.fetchall():
                key = f"{row['name'].lower()}|{row['address'].lower()}"
                seen_global.add(key)
        logger.info("Deduplication index populated from SQLite. Currently has %d keys.", len(seen_global))
    except Exception as e:
        logger.error("Failed to load leads from SQLite database: %s", e)

    # 2. From Excel (for backwards compatibility/safety)
    import os
    if os.path.exists(args.output):
        try:
            logger.info("Loading existing leads from %s for deduplication ...", args.output)
            df_existing = pd.read_excel(args.output)
            for _, row in df_existing.iterrows():
                key = f"{str(row.get('Name', '')).lower()}|{str(row.get('Address', '')).lower()}"
                seen_global.add(key)
            logger.info("Total deduplication index has %d keys (SQLite + Excel).", len(seen_global))
        except Exception as e:
            logger.warning("Failed to load existing leads from Excel: %s", e)

    with sync_playwright() as playwright:
        browser, context, page = initialize_browser(playwright)

        try:
            for query in queries:
                logger.info("=" * 60)
                logger.info("Starting query: '%s'  (max %d per query, %d total so far)",
                            query, args.max, len(all_records))
                logger.info("=" * 60)

                success = perform_search(page, query)
                if not success:
                    logger.error("Search failed for query: '%s' - skipping.", query)
                    continue

                records = extract_business_data(page, args.max)

                # Cross-query deduplication
                for record in records:
                    key = f"{record['Name'].lower()}|{record['Address'].lower()}"
                    if key not in seen_global:
                        seen_global.add(key)
                        all_records.append(record)

                logger.info("Query done. Cumulative unique records: %d", len(all_records))
                human_delay(3, 6)  # polite pause between queries

        except KeyboardInterrupt:
            logger.warning("Interrupted by user - saving collected data before exit ...")
        except Exception as exc:
            logger.error("Unexpected error: %s", exc, exc_info=True)
        finally:
            context.close()
            browser.close()
            logger.info("Browser closed.")

    # ── Check A: Name frequency filter ───────────────────────────────────────
    # Any business name appearing 3+ times across all queries is likely a chain
    # (e.g. "Sharma's Cafe" with 4 outlets in different Delhi areas)
    from collections import Counter
    name_freq = Counter(r["Name"].strip().lower() for r in all_records)
    chain_names = {name for name, count in name_freq.items() if count >= 3}
    if chain_names:
        before = len(all_records)
        all_records = [r for r in all_records
                       if r["Name"].strip().lower() not in chain_names]
        removed = before - len(all_records)
        logger.info(
            "Name-frequency filter removed %d records (%d unique chain names detected)",
            removed, len(chain_names)
        )
        for cn in sorted(chain_names):
            logger.debug("  Flagged as chain by frequency: '%s' (appeared %d×)",
                         cn, name_freq[cn])

    if all_records:
        save_to_excel(all_records, args.output)
        print(f"\n[DONE] Done! {len(all_records)} leads saved to '{args.output}'")
    else:
        logger.warning("No records collected. Excel file was not created.")
        print("\n[WARNING]  No records collected. Please check your internet connection or selectors.")


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    main()
