"""
filter_leads.py — Post-processing script to remove big brands/malls
from an already-collected delhi_leads.xlsx file.

Run this AFTER scraper.py finishes:
    python filter_leads.py
    python filter_leads.py --input delhi_leads.xlsx --output delhi_leads_clean.xlsx
"""

import sys
# Set stdout/stderr encoding to utf-8 to prevent charmap crashes on Windows
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')

import argparse
import pandas as pd
import config

def _is_platform_website(url: str) -> bool:
    """Check if the URL belongs to a known platform."""
    if not isinstance(url, str) or url == "N/A":
        return False
    url_lower = url.lower()
    return any(domain in url_lower for domain in config.PLATFORM_DOMAINS)


def _score_lead(website: str, rating: float) -> str:
    """Assign priority based on website presence and rating."""
    if not isinstance(website, str) or website == "N/A":
        return "[HIGH] High"
    if _is_platform_website(website):
        if rating and rating >= config.HIGH_RATING_THRESHOLD:
            return "[MEDIUM] Medium"
        return "[LOW] Low"
    return None # Signals this should be filtered out


def is_local_business(name: str, reviews, website, rating, phone="N/A") -> bool:
    """Return True if the listing is a small/local business and doesn't have a personal website."""
    if not isinstance(name, str):
        return True

    name_lower = name.lower()

    # Check brand blacklist
    for brand in config.BRAND_BLACKLIST:
        if brand.lower() in name_lower:
            return False

    # Check review count cap
    try:
        rev = int(reviews)
        if rev > config.MAX_REVIEWS_FOR_LOCAL:
            return False
    except (ValueError, TypeError):
        pass

    # NEW: Strict Personal Website Filter
    if isinstance(website, str) and website != "N/A":
        if not _is_platform_website(website):
            return False

    # NEW: Contact Details (Phone) Requirement Filter
    if getattr(config, 'REQUIRE_PHONE_NUMBER', False):
        if not isinstance(phone, str) or phone == "N/A":
            return False

    return True


def filter_excel(input_file: str, output_file: str) -> None:
    print(f"[READING] Reading: {input_file}")
    df = pd.read_excel(input_file)
    total_before = len(df)
    print(f"   Total rows before filter: {total_before}")

    # Apply filter
    mask = df.apply(
        lambda row: is_local_business(
            row.get("Name", ""), 
            row.get("Total Reviews"), 
            row.get("Website", "N/A"),
            row.get("Rating"),
            phone=row.get("Phone", "N/A")
        ),
        axis=1
    )
    df_clean = df[mask].copy()

    # Re-calculate Priority
    df_clean["Priority"] = df_clean.apply(
        lambda row: _score_lead(row.get("Website", "N/A"), row.get("Rating")),
        axis=1
    )

    # Helper to identify locality
    def get_locality(address):
        if not isinstance(address, str):
            return "other"
        addr_lower = address.lower()
        localities = [
            "hauz khas", "dwarka", "rajouri", "connaught", "lajpat", 
            "karol bagh", "saket", "uttam nagar", "shahpur jat", "janakpuri", "subhash nagar"
        ]
        for loc in localities:
            if loc in addr_lower:
                return loc
        return "other"

    # Locality clean names map
    locality_names = {
        "hauz khas": "Hauz Khas",
        "dwarka": "Dwarka",
        "rajouri": "Rajouri Garden",
        "connaught": "Connaught Place",
        "lajpat": "Lajpat Nagar",
        "karol bagh": "Karol Bagh",
        "saket": "Saket",
        "uttam nagar": "Uttam Nagar",
        "shahpur jat": "Shahpur Jat",
        "janakpuri": "Janakpuri",
        "subhash nagar": "Subhash Nagar",
        "other": "Other"
    }

    # Add Locality column
    df_clean["Locality"] = df_clean["Address"].apply(get_locality).map(locality_names).fillna("Other")

    # Add locality helper for grouping
    df_clean["_locality"] = df_clean["Address"].apply(get_locality)
    
    # ── Locality & Priority Sorting ──────────────────────────────────
    # Sort: High (1) -> Medium (2) -> Low (3)
    priority_map = {"[HIGH] High": 1, "[MEDIUM] Medium": 2, "[LOW] Low": 3}
    df_clean["_priority_sort"] = df_clean["Priority"].map(priority_map).fillna(99)
    
    # Sort by locality first (so same location is grouped), then priority, then name
    df_clean = df_clean.sort_values(by=["_locality", "_priority_sort", "Name"]).drop(columns=["_priority_sort", "_locality"]).reset_index(drop=True)

    # Reorder columns to place Locality right after Address
    cols = [
        "Name", "Rating", "Total Reviews", "Address", "Locality", "Phone", 
        "Website", "Category", "Maps URL", "Priority", "Competitor Name", "Competitor Website"
    ]
    df_clean = df_clean[cols]

    removed = total_before - len(df_clean)

    print(f"   Removed (brands/malls/chains): {removed}")
    print(f"[SUCCESS] Remaining local business leads (grouped by Locality): {len(df_clean)}")

    # Save cleaned file
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df_clean.to_excel(writer, index=False, sheet_name="Delhi Leads")
        ws = writer.sheets["Delhi Leads"]

        from openpyxl.styles import Font, PatternFill, Alignment

        # Define 5 slightly darker, rich but warm pastel colors (eye-friendly)
        fills = {
            "blue": PatternFill(start_color="C9DAF8", end_color="C9DAF8", fill_type="solid"),
            "pink": PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
            "peach": PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid"),
            "cream": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            "green": PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
            "white": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        }

        # Dynamically map unique localities to the 5 soft colors
        unique_localities = []
        for addr in df_clean["Address"]:
            loc = get_locality(addr)
            if loc not in unique_localities and loc != "other":
                unique_localities.append(loc)

        color_order = ["blue", "pink", "peach", "cream", "green"]
        locality_colors = {}
        for idx, loc in enumerate(unique_localities):
            color_name = color_order[idx % len(color_order)]
            locality_colors[loc] = fills[color_name]

        # Apply pastel row fills (excluding header)
        for r_idx, row_cells in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            addr = df_clean.iloc[r_idx - 2]["Address"]
            loc = get_locality(addr)
            fill_to_apply = locality_colors.get(loc, fills["white"])
            for cell in row_cells:
                cell.fill = fill_to_apply

        # Style header row
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        for col_cells in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A2"

    print(f"[SUCCESS] Saved clean file: {output_file}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Filter big brands from leads Excel")
    parser.add_argument("--input",  default="delhi_leads.xlsx",       help="Input Excel file")
    parser.add_argument("--output", default="delhi_leads_clean.xlsx", help="Output cleaned file")
    args = parser.parse_args()
    filter_excel(args.input, args.output)
